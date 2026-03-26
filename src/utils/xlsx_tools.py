import json
from pathlib import Path
from typing import List

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Цвета статусов (idStatus → hex fill)
STATUS_COLORS = {
    6: "C6EFCE",  # Действует         → зелёный
    1: "D9D9D9",  # Архивный          → серый
    14: "FFC7CE",  # Прекращен         → красный
    15: "FFEB9C",  # Приостановлен     → оранжевый
    19: "FFEB9C",  # Частично приост.  → жёлтый
}

COLUMNS = [
    ("Название предприятия (Заявитель)", "applicant_full_name"),
    ("ИНН", "applicant_inn"),
    ("Можно ли работать?", "_can_work"),
    ("Дата стоп-листа", "_stop_date"),
    ("Статус", "name_status"),
    ("Другие статусы", "_other_statuses"),
    ("За кем закреплено", "fa_name"),
    ("Адрес", "address"),
    ("Тип аккредитованного лица", "name_type"),
    ("Номер записи в РАЛ", "reg_number"),
    ("Телефоны", "phones"),
    ("Email", "emails"),
    ("ФИО руководителя", "head_person_fio"),
]

THIN = Side(style="thin")


def _border():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _header_fill():
    return PatternFill(fill_type="solid", fgColor="4472C4")


def _status_fill(id_status: int) -> PatternFill:
    color = STATUS_COLORS.get(id_status, "FFFFFF")
    return PatternFill(fill_type="solid", fgColor=color)


def _can_work(id_status) -> str:
    return "Да" if id_status == 6 else "Нет"


def _extract_value(record: dict, field: str) -> str:
    if field == "_can_work":
        return _can_work(record.get("id_status"))
    if field == "_stop_date":
        # Дата стоп-листа — берём reg_date если статус не "Действует"
        if record.get("id_status") != 6:
            return record.get("reg_date", "") or ""
        return ""
    if field == "_other_statuses":
        return ""
    val = record.get(field)
    if val is None:
        return ""
    # phones / emails хранятся как JSON-массивы в БД
    if field in ("phones", "emails"):
        if isinstance(val, str):
            try:
                val = json.loads(val)
            except Exception:
                return val
        if isinstance(val, list):
            return ", ".join(str(v) for v in val)
    return str(val)


class XLSXExporter:
    def __init__(self, output_path: str = "data/export.xlsx"):
        self.output_path = Path(output_path)
        self.output_path.parent.mkdir(parents=True, exist_ok=True)

    def export(self, records: List[dict]) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "РАЛ"

        # Заголовки
        headers = [col[0] for col in COLUMNS]
        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = _header_fill()
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = _border()

        ws.row_dimensions[1].height = 40

        # Данные
        for row_idx, record in enumerate(records, start=2):
            id_status = record.get("id_status")
            fill = _status_fill(id_status)

            for col_idx, (_, field) in enumerate(COLUMNS, start=1):
                value = _extract_value(record, field)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = _border()

        # Ширина столбцов
        col_widths = [40, 15, 18, 15, 20, 18, 25, 40, 25, 20, 25, 30, 30]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        wb.save(self.output_path)
        logger.info(f"Exported {len(records)} records to {self.output_path}")
        return self.output_path

    def _create_workbook(self, records: List[dict]) -> Workbook:
        """Create workbook for download (without saving to file)."""
        wb = Workbook()
        ws = wb.active
        ws.title = "РАЛ"

        # Headers
        headers = [col[0] for col in COLUMNS]
        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = _header_fill()
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = _border()

        ws.row_dimensions[1].height = 40

        # Data
        for row_idx, record in enumerate(records, start=2):
            id_status = record.get("id_status")
            fill = _status_fill(id_status)

            for col_idx, (_, field) in enumerate(COLUMNS, start=1):
                value = _extract_value(record, field)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = _border()

        # Column widths
        col_widths = [40, 15, 18, 15, 20, 18, 25, 40, 25, 20, 25, 30, 30]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        return wb
